from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from backend.core.refund_detection import StatementParserFactory
from backend.core.refund_detection.models import NormalizedTransaction
from backend.core.refund_detection.utils import EMAIL_REGEX, extract_email, is_multiple_of, normalize_text


DEFAULT_REFUND_PHRASES = [
    "vozvrat oplaty po dogovoru",
    "возврат оплаты по договору",
    "возврат оплат по договору",
    "возврат по договору",
]


class RefundReportService:
    def __init__(self):
        self.parser_factory = StatementParserFactory()

    def analyze_files(self, file_infos: List[Dict[str, Any]], options: Dict[str, Any]) -> Dict[str, Any]:
        applied_options = self._normalize_options(options)
        transactions: List[NormalizedTransaction] = []
        warnings: List[str] = []
        by_file_totals: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"bank": "", "matched_transactions": 0, "matched_amount": 0.0})
        scanned_transactions = 0

        for file_info in file_infos:
            original_name = file_info["original_name"]
            try:
                parser = self.parser_factory.get_parser(file_info["temp_path"], original_name)
                parsed_transactions = parser.parse(file_info["temp_path"], original_name)
            except Exception as exc:
                warnings.append(f"{original_name}: {exc}")
                continue

            for transaction in parsed_transactions:
                if applied_options["outgoing_only"] and transaction.direction != "outgoing":
                    continue

                scanned_transactions += 1
                matched_rules = self._evaluate_rules(transaction, applied_options)
                enabled_rules_count = self._enabled_rules_count(applied_options)
                is_match = self._is_match(matched_rules, enabled_rules_count, applied_options)
                if not is_match:
                    continue

                transaction.extracted_email = extract_email(transaction.payment_purpose)
                transaction.matched_rules = matched_rules
                transactions.append(transaction)

                file_summary = by_file_totals[transaction.source_file]
                file_summary["bank"] = transaction.source_bank
                file_summary["matched_transactions"] += 1
                file_summary["matched_amount"] += float(transaction.amount or 0)

        transactions.sort(key=lambda item: (item.operation_date, item.source_file, item.document_number))
        response_transactions = [transaction.to_dict() for transaction in transactions]
        summary = self._build_summary(file_infos, transactions, by_file_totals, warnings, scanned_transactions)

        return {
            "success": True,
            "summary": summary,
            "transactions": response_transactions,
            "warnings": warnings,
            "applied_options": applied_options,
        }

    def export_report(self, analysis: Dict[str, Any]) -> bytes:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        summary = analysis.get("summary", {})
        applied_options = analysis.get("applied_options", {})
        transactions = analysis.get("transactions", [])

        self._write_kv_sheet(
            summary_sheet,
            [
                ("Generated at", summary.get("generated_at", "")),
                ("Requested files", summary.get("requested_files", 0)),
                ("Processed files", summary.get("processed_files", 0)),
                ("Skipped files", summary.get("skipped_files", 0)),
                ("Scanned transactions", summary.get("total_transactions_scanned", 0)),
                ("Matched transactions", summary.get("matched_transactions", 0)),
                ("Matched total amount", summary.get("total_amount", 0)),
                ("Unique emails", summary.get("unique_emails_count", 0)),
                ("Match mode", applied_options.get("match_mode", "any")),
                ("Outgoing only", applied_options.get("outgoing_only", True)),
                ("Amount multiple enabled", applied_options.get("enable_amount_multiple", True)),
                ("Amount multiple", applied_options.get("amount_multiple", 5000)),
                ("Email rule enabled", applied_options.get("enable_email", True)),
                ("Refund phrase rule enabled", applied_options.get("enable_refund_phrase", True)),
                ("Refund phrases", ", ".join(applied_options.get("refund_phrases", []))),
            ],
        )

        matched_sheet = workbook.create_sheet("Matched Transactions")
        matched_headers = [
            "Operation date",
            "Amount",
            "Direction",
            "Bank",
            "Account",
            "Document",
            "Counterparty",
            "Email",
            "Matched rules",
            "Source file",
            "Payment purpose",
        ]
        matched_rows = [
            [
                row.get("operation_date", ""),
                row.get("amount", 0),
                row.get("direction", ""),
                row.get("source_bank", ""),
                row.get("account_number", ""),
                row.get("document_number", ""),
                row.get("counterparty_name", ""),
                row.get("extracted_email", ""),
                ", ".join(row.get("matched_rules", [])),
                row.get("source_file", ""),
                row.get("payment_purpose", ""),
            ]
            for row in transactions
        ]
        self._write_table(matched_sheet, matched_headers, matched_rows)

        emails_sheet = workbook.create_sheet("Unique Emails")
        email_rows = [[email] for email in summary.get("unique_emails", [])]
        self._write_table(emails_sheet, ["Email"], email_rows)

        files_sheet = workbook.create_sheet("By File")
        file_headers = ["Source file", "Bank", "Matched transactions", "Matched amount"]
        file_rows = [
            [
                row.get("source_file", ""),
                row.get("source_bank", ""),
                row.get("matched_transactions", 0),
                row.get("matched_amount", 0),
            ]
            for row in summary.get("by_file", [])
        ]
        self._write_table(files_sheet, file_headers, file_rows)

        if analysis.get("warnings"):
            warnings_sheet = workbook.create_sheet("Warnings")
            self._write_table(warnings_sheet, ["Warning"], [[warning] for warning in analysis["warnings"]])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _normalize_options(self, options: Dict[str, Any]) -> Dict[str, Any]:
        phrases = options.get("refund_phrases") or DEFAULT_REFUND_PHRASES
        cleaned_phrases = []
        for phrase in phrases:
            cleaned = " ".join(str(phrase).strip().split())
            if cleaned:
                cleaned_phrases.append(cleaned)

        return {
            "enable_amount_multiple": bool(options.get("enable_amount_multiple", True)),
            "amount_multiple": float(options.get("amount_multiple", 5000) or 0),
            "enable_email": bool(options.get("enable_email", True)),
            "enable_refund_phrase": bool(options.get("enable_refund_phrase", True)),
            "refund_phrases": cleaned_phrases or DEFAULT_REFUND_PHRASES,
            "match_mode": "all" if options.get("match_mode") == "all" else "any",
            "outgoing_only": bool(options.get("outgoing_only", True)),
        }

    def _enabled_rules_count(self, options: Dict[str, Any]) -> int:
        return sum(
            [
                1 if options["enable_amount_multiple"] and options["amount_multiple"] > 0 else 0,
                1 if options["enable_email"] else 0,
                1 if options["enable_refund_phrase"] else 0,
            ]
        )

    def _is_match(self, matched_rules: List[str], enabled_rules_count: int, options: Dict[str, Any]) -> bool:
        if enabled_rules_count == 0:
            return False
        if options["match_mode"] == "all":
            return len(matched_rules) == enabled_rules_count

        if len(matched_rules) == 0:
            return False

        semantic_rules_enabled = []
        if options["enable_email"]:
            semantic_rules_enabled.append("email")
        if options["enable_refund_phrase"]:
            semantic_rules_enabled.append("refund_phrase")

        if semantic_rules_enabled:
            return any(rule in matched_rules for rule in semantic_rules_enabled)

        return len(matched_rules) > 0

    def _evaluate_rules(self, transaction: NormalizedTransaction, options: Dict[str, Any]) -> List[str]:
        matched_rules: List[str] = []
        purpose_normalized = normalize_text(transaction.payment_purpose)

        if options["enable_amount_multiple"] and options["amount_multiple"] > 0 and is_multiple_of(transaction.amount, options["amount_multiple"]):
            matched_rules.append("amount_multiple")

        if options["enable_email"] and EMAIL_REGEX.search(transaction.payment_purpose or ""):
            matched_rules.append("email")

        if options["enable_refund_phrase"] and self._matches_refund_phrase(purpose_normalized, options["refund_phrases"]):
            matched_rules.append("refund_phrase")

        return matched_rules

    def _matches_refund_phrase(self, normalized_purpose: str, phrases: List[str]) -> bool:
        if not normalized_purpose:
            return False
        phrase_match = any(normalize_text(phrase) in normalized_purpose for phrase in phrases)
        heuristic_match = (
            "возврат" in normalized_purpose
            and "договор" in normalized_purpose
            and ("оплат" in normalized_purpose or "платеж" in normalized_purpose)
        )
        return phrase_match or heuristic_match

    def _build_summary(
        self,
        file_infos: List[Dict[str, Any]],
        transactions: List[NormalizedTransaction],
        by_file_totals: Dict[str, Dict[str, Any]],
        warnings: List[str],
        scanned_transactions: int,
    ) -> Dict[str, Any]:
        by_bank: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"source_bank": "", "matched_transactions": 0, "matched_amount": 0.0})
        unique_emails = sorted({transaction.extracted_email for transaction in transactions if transaction.extracted_email})

        for transaction in transactions:
            bank_summary = by_bank[transaction.source_bank]
            bank_summary["source_bank"] = transaction.source_bank
            bank_summary["matched_transactions"] += 1
            bank_summary["matched_amount"] += float(transaction.amount or 0)

        return {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "requested_files": len(file_infos),
            "processed_files": len(file_infos) - len(warnings),
            "skipped_files": len(warnings),
            "total_transactions_scanned": scanned_transactions,
            "matched_transactions": len(transactions),
            "total_amount": round(sum(float(transaction.amount or 0) for transaction in transactions), 2),
            "unique_emails_count": len(unique_emails),
            "unique_emails": unique_emails,
            "banks": sorted(by_bank.keys()),
            "by_bank": [
                {
                    "source_bank": item["source_bank"],
                    "matched_transactions": item["matched_transactions"],
                    "matched_amount": round(item["matched_amount"], 2),
                }
                for item in sorted(by_bank.values(), key=lambda value: value["source_bank"])
            ],
            "by_file": [
                {
                    "source_file": source_file,
                    "source_bank": value["bank"],
                    "matched_transactions": value["matched_transactions"],
                    "matched_amount": round(value["matched_amount"], 2),
                }
                for source_file, value in sorted(by_file_totals.items(), key=lambda item: item[0])
            ],
        }

    def _write_kv_sheet(self, sheet, rows: List[Any]) -> None:
        header_fill = PatternFill("solid", fgColor="1F7A4C")
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)
            sheet.cell(row=row_index, column=1).font = Font(bold=True)
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 80
        sheet["A1"].fill = header_fill
        sheet["A1"].font = Font(bold=True, color="FFFFFF")

    def _write_table(self, sheet, headers: List[str], rows: List[List[Any]]) -> None:
        header_fill = PatternFill("solid", fgColor="1F7A4C")
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        for row_index, row in enumerate(rows, start=2):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
