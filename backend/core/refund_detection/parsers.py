from __future__ import annotations

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from .models import NormalizedTransaction
from .utils import (
    clean_account_number,
    find_row_by_keywords,
    has_meaningful_values,
    parse_amount,
    parse_date,
    read_csv_rows,
)


class BaseStatementParser(ABC):
    parser_name = "base"
    bank_name = "unknown"

    @abstractmethod
    def can_handle(self, file_path: str, original_name: str) -> bool:
        raise NotImplementedError

    @abstractmethod
    def parse(self, file_path: str, original_name: str) -> List[NormalizedTransaction]:
        raise NotImplementedError


class SberBusinessParser(BaseStatementParser):
    parser_name = "sber_business"
    bank_name = "sber_business"

    def can_handle(self, file_path: str, original_name: str) -> bool:
        if "сбербизнес" in original_name.lower():
            return True
        if Path(file_path).suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            return False

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                header_row = find_row_by_keywords(
                    worksheet.iter_rows(min_row=1, max_row=20, values_only=True),
                    ["дата проводки", "сумма по дебету", "сумма по кредиту", "назначение платежа"],
                )
                if header_row:
                    return True
        finally:
            workbook.close()
        return False

    def parse(self, file_path: str, original_name: str) -> List[NormalizedTransaction]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        transactions: List[NormalizedTransaction] = []
        try:
            worksheet, header_row = self._resolve_sheet(workbook)
            if worksheet is None or header_row is None:
                raise ValueError(f"Could not locate SberBusiness table in {original_name}")

            headers = [str(cell or "").strip() for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
            indexes = self._map_indexes(headers)
            account_number = clean_account_number(worksheet.title)

            for row in worksheet.iter_rows(min_row=header_row + 2, values_only=True):
                if not has_meaningful_values(list(row)):
                    continue

                date_value = row[indexes["date"]] if indexes["date"] < len(row) else None
                debit_value = parse_amount(row[indexes["debit"]] if indexes["debit"] < len(row) else None)
                credit_value = parse_amount(row[indexes["credit"]] if indexes["credit"] < len(row) else None)
                purpose_value = str(row[indexes["purpose"]] if indexes["purpose"] < len(row) and row[indexes["purpose"]] is not None else "").strip()
                document_number = str(row[indexes["document"]] if indexes["document"] < len(row) and row[indexes["document"]] is not None else "").strip()

                if not str(date_value or "").strip() and not purpose_value and not debit_value and not credit_value:
                    continue

                direction = "outgoing" if debit_value > 0 else "incoming" if credit_value > 0 else "unknown"
                amount = debit_value if debit_value > 0 else credit_value

                transactions.append(
                    NormalizedTransaction(
                        source_bank=self.bank_name,
                        source_parser=self.parser_name,
                        source_file=original_name,
                        account_number=account_number,
                        operation_date=parse_date(date_value),
                        document_number=document_number,
                        direction=direction,
                        amount=amount,
                        counterparty_name="",
                        payment_purpose=purpose_value,
                    )
                )
        finally:
            workbook.close()
        return transactions

    def _resolve_sheet(self, workbook) -> Tuple[Optional[object], Optional[int]]:
        for worksheet in workbook.worksheets:
            header_row = find_row_by_keywords(
                worksheet.iter_rows(min_row=1, max_row=20, values_only=True),
                ["дата проводки", "сумма по дебету", "сумма по кредиту", "назначение платежа"],
            )
            if header_row:
                return worksheet, header_row
        return None, None

    def _map_indexes(self, headers: Sequence[str]) -> Dict[str, int]:
        mapping = {"date": None, "debit": None, "credit": None, "purpose": None, "document": None}
        for index, header in enumerate(headers):
            normalized = header.lower().strip()
            if "дата проводки" in normalized and mapping["date"] is None:
                mapping["date"] = index
            elif "сумма по дебету" in normalized and mapping["debit"] is None:
                mapping["debit"] = index
            elif "сумма по кредиту" in normalized and mapping["credit"] is None:
                mapping["credit"] = index
            elif "назначение платежа" in normalized and mapping["purpose"] is None:
                mapping["purpose"] = index
            elif "№ документа" in normalized and mapping["document"] is None:
                mapping["document"] = index

        missing = [key for key, value in mapping.items() if value is None and key != "document"]
        if missing:
            raise ValueError(f"SberBusiness header mapping failed, missing: {', '.join(missing)}")
        mapping["document"] = mapping["document"] if mapping["document"] is not None else 0
        return mapping


class AccountStatementParser(BaseStatementParser):
    parser_name = "account_statement"
    bank_name = "account_statement"

    def can_handle(self, file_path: str, original_name: str) -> bool:
        if "выписка_" in original_name.lower():
            return True
        if Path(file_path).suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            return False

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                header_row = find_row_by_keywords(
                    worksheet.iter_rows(min_row=1, max_row=25, values_only=True),
                    ["дата", "дебет", "кредит", "назначение платежа"],
                )
                if header_row:
                    return True
        finally:
            workbook.close()
        return False

    def parse(self, file_path: str, original_name: str) -> List[NormalizedTransaction]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        transactions: List[NormalizedTransaction] = []
        try:
            worksheet, header_row = self._resolve_sheet(workbook)
            if worksheet is None or header_row is None:
                raise ValueError(f"Could not locate account statement table in {original_name}")

            top_headers = [str(cell or "").strip() for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
            sub_headers = [str(cell or "").strip() for cell in next(worksheet.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))]
            headers = self._combine_headers(top_headers, sub_headers)
            indexes = self._map_indexes(headers)
            account_number = self._extract_account_number(worksheet)

            for row in worksheet.iter_rows(min_row=header_row + 2, values_only=True):
                if not has_meaningful_values(list(row)):
                    continue

                date_value = row[indexes["date"]] if indexes["date"] < len(row) else None
                debit_value = parse_amount(row[indexes["debit"]] if indexes["debit"] < len(row) else None)
                credit_value = parse_amount(row[indexes["credit"]] if indexes["credit"] < len(row) else None)
                purpose_value = str(row[indexes["purpose"]] if indexes["purpose"] < len(row) and row[indexes["purpose"]] is not None else "").strip()
                document_number = str(row[indexes["document"]] if indexes["document"] < len(row) and row[indexes["document"]] is not None else "").strip()
                counterparty_name = str(row[indexes["counterparty"]] if indexes["counterparty"] < len(row) and row[indexes["counterparty"]] is not None else "").strip()

                if not str(date_value or "").strip() and not purpose_value and not debit_value and not credit_value:
                    continue

                direction = "outgoing" if debit_value > 0 else "incoming" if credit_value > 0 else "unknown"
                amount = debit_value if debit_value > 0 else credit_value

                transactions.append(
                    NormalizedTransaction(
                        source_bank=self.bank_name,
                        source_parser=self.parser_name,
                        source_file=original_name,
                        account_number=account_number,
                        operation_date=parse_date(date_value),
                        document_number=document_number,
                        direction=direction,
                        amount=amount,
                        counterparty_name=counterparty_name,
                        payment_purpose=purpose_value,
                    )
                )
        finally:
            workbook.close()
        return transactions

    def _resolve_sheet(self, workbook) -> Tuple[Optional[object], Optional[int]]:
        for worksheet in workbook.worksheets:
            header_row = find_row_by_keywords(
                worksheet.iter_rows(min_row=1, max_row=25, values_only=True),
                ["дата", "дебет", "кредит", "назначение платежа"],
            )
            if header_row:
                return worksheet, header_row
        return None, None

    def _combine_headers(self, top_headers: Sequence[str], sub_headers: Sequence[str]) -> List[str]:
        combined = []
        for top, sub in zip(top_headers, sub_headers):
            top_clean = top.strip()
            sub_clean = sub.strip()
            combined.append(f"{top_clean} {sub_clean}".strip())
        return combined

    def _map_indexes(self, headers: Sequence[str]) -> Dict[str, int]:
        mapping = {"date": None, "debit": None, "credit": None, "purpose": None, "document": None, "counterparty": None}
        for index, header in enumerate(headers):
            normalized = header.lower().strip()
            if normalized == "дата" and mapping["date"] is None:
                mapping["date"] = index
            elif normalized == "дебет" and mapping["debit"] is None:
                mapping["debit"] = index
            elif normalized == "кредит" and mapping["credit"] is None:
                mapping["credit"] = index
            elif "назначение платежа" in normalized and mapping["purpose"] is None:
                mapping["purpose"] = index
            elif "номер документа" in normalized and mapping["document"] is None:
                mapping["document"] = index
            elif "контрагент наименование" in normalized and mapping["counterparty"] is None:
                mapping["counterparty"] = index

        missing = [key for key, value in mapping.items() if value is None and key != "counterparty"]
        if missing:
            raise ValueError(f"Account statement header mapping failed, missing: {', '.join(missing)}")
        mapping["counterparty"] = mapping["counterparty"] if mapping["counterparty"] is not None else 0
        mapping["document"] = mapping["document"] if mapping["document"] is not None else 0
        return mapping

    def _extract_account_number(self, worksheet) -> str:
        for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                digits = clean_account_number(cell)
                if len(digits) >= 20:
                    return digits
        return ""


class MigTorgCsvParser(BaseStatementParser):
    parser_name = "migtorg_csv"
    bank_name = "migtorg"

    def can_handle(self, file_path: str, original_name: str) -> bool:
        lower_name = original_name.lower()
        if "migtorg" in lower_name:
            return True
        if Path(file_path).suffix.lower() != ".csv":
            return False
        try:
            dataframe = read_csv_rows(file_path)
        except Exception:
            return False
        headers = {header.lower().strip() for header in dataframe.columns}
        return "operation_type" in headers and "amount / amount" in headers

    def parse(self, file_path: str, original_name: str) -> List[NormalizedTransaction]:
        dataframe = read_csv_rows(file_path)
        transactions: List[NormalizedTransaction] = []
        for _, row in dataframe.fillna("").iterrows():
            raw_amount = parse_amount(row.get("amount / amount", 0))
            currency = str(row.get("currency / currency", "") or row.get("real_currency / channel_currency", "")).strip().upper()
            amount = raw_amount / 100 if currency == "RUB" and raw_amount >= 1000 else raw_amount
            operation_type = str(row.get("operation_type", "")).strip()
            direction = "outgoing" if any(token in operation_type.lower() for token in ["refund", "voucher", "payout"]) else "incoming"
            purpose = " | ".join(
                part for part in [
                    str(row.get("operation_type", "")).strip(),
                    str(row.get("operation_status", "")).strip(),
                    str(row.get("card_holder", "")).strip(),
                ] if part
            )

            transactions.append(
                NormalizedTransaction(
                    source_bank=self.bank_name,
                    source_parser=self.parser_name,
                    source_file=original_name,
                    account_number=str(row.get("customer_purse / account_number", "")).strip(),
                    operation_date=parse_date(row.get("completed_at / operation_completed_at", "")),
                    document_number=str(row.get("id / operation_id", "") or row.get("external_id / payment_id", "")).strip(),
                    direction=direction,
                    amount=amount,
                    counterparty_name=str(row.get("card_holder", "")).strip(),
                    payment_purpose=purpose,
                )
            )
        return transactions


class StatementParserFactory:
    def __init__(self):
        self.parsers = [
            SberBusinessParser(),
            AccountStatementParser(),
            MigTorgCsvParser(),
        ]

    def get_parser(self, file_path: str, original_name: str) -> BaseStatementParser:
        for parser in self.parsers:
            if parser.can_handle(file_path, original_name):
                return parser
        raise ValueError(f"Unsupported statement format: {original_name}")
